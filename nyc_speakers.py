#!/usr/bin/env python3
"""
NFT NYC Speaker Scraper - All Tracks Version (Professional Output) - OPTIMIZED
⚡ Extracts speakers in EXACT website order with professional formatting
"""

import asyncio
import logging
import re
import json
import argparse
from datetime import datetime
from typing import List, Dict, Set
from time import time

import pandas as pd
from playwright.async_api import async_playwright, Page, Locator
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- Configuration ---
BASE_URL = "https://www.nft.nyc"
BASE_FILENAME = "nyc_speaker_all_tracks"
USE_TIMESTAMP = True

# Track definitions: (Display Name, URL Path)
TRACKS = [
    ("FEATURED", "/speakers"),
    ("COMMUNITY", "/speakers/community"),
    ("AI", "/speakers/ai"),
    ("ART", "/speakers/art"),
    ("ENTERTAINMENT", "/speakers/entertainment"),
    ("LEGAL", "/speakers/legal"),
    ("BRANDS", "/speakers/brands"),
    ("FUTURE", "/speakers/future"),
    ("GAMING", "/speakers/gaming"),
    ("BTC & ORDINALS", "/speakers/bitcoin"),
]

# ⚡ PERFORMANCE OPTIMIZATION SETTINGS
MODAL_WAIT_MS = 700
MODAL_CLOSE_MS = 350
TRACK_PAUSE_MS = 500
PAGE_LOAD_WAIT_MS = 3000  # Wait for all content to render

# --- Logging ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%H:%M:%S')
logger = logging.getLogger(__name__)


class SocialMediaExtractor:
    """Extract social media handles."""

    INVALID_X = {'home', 'explore', 'i', 'intent', 'share', 'login', 'uwt', 'nftnyc', 'twitter'}
    INVALID_IG = {'explore', 'accounts', 'direct', 'uwt', 'nftnyc', 'instagram'}
    INVALID_LI = {'feed', 'jobs', 'help', 'about', 'linkedin'}

    @classmethod
    def extract_x(cls, text: str) -> str:
        if not text:
            return "N/A"
        match = re.search(r'(?:twitter\.com|x\.com)/[@]?([A-Za-z0-9_]{1,15})', text, re.I)
        if match:
            handle = match.group(1).lower()
            if handle not in cls.INVALID_X and not handle.isdigit():
                return handle
        return "N/A"

    @classmethod
    def extract_instagram(cls, text: str) -> str:
        if not text:
            return "N/A"
        match = re.search(r'instagram\.com/([A-Za-z0-9_.]{1,30})', text, re.I)
        if match:
            handle = match.group(1).lower()
            if handle not in cls.INVALID_IG and not (handle.startswith('.') or handle.endswith('.')):
                return handle
        return "N/A"

    @classmethod
    def extract_linkedin(cls, text: str) -> str:
        if not text:
            return "N/A"
        match = re.search(r'linkedin\.com/(in|company)/([A-Za-z0-9-]{3,100})', text, re.I)
        if match:
            path_type, path = match.groups()
            if path.lower() not in cls.INVALID_LI:
                return f"linkedin.com/{path_type}/{path}"
        return "N/A"


async def dismiss_cookie_banner(page: Page):
    """Dismiss cookie consent banner if present."""
    try:
        cookie_banner = page.locator('#hs-eu-cookie-confirmation').first
        if await cookie_banner.is_visible(timeout=1000):
            accept_btn = cookie_banner.locator('button, a').first
            if await accept_btn.count() > 0:
                await accept_btn.click()
                await page.wait_for_timeout(300)
    except:
        pass


async def close_modal_completely(page: Page) -> bool:
    """Close modal by clicking overlay."""
    try:
        close_btn = page.locator('button.sz-modal__close').first
        if await close_btn.is_visible(timeout=600):
            await close_btn.click()
            await page.wait_for_timeout(MODAL_CLOSE_MS)
            return True

        overlay = page.locator('.sz-modal-overlay').first
        if await overlay.is_visible(timeout=400):
            await overlay.click()
            await page.wait_for_timeout(MODAL_CLOSE_MS)

        await page.keyboard.press('Escape')
        await page.wait_for_timeout(200)

        return True
    except:
        return False


async def extract_social_from_modal(modal: Locator) -> Dict[str, str]:
    """Extract social media from modal."""
    social = {"x_handle": "N/A", "instagram": "N/A", "linkedin": "N/A"}

    try:
        social_list = modal.locator('ul.sz-speaker__links')
        if await social_list.count() > 0:
            links = social_list.locator('a[href]')
            count = await links.count()

            for i in range(count):
                try:
                    href = await links.nth(i).get_attribute('href', timeout=600)
                    if not href:
                        continue

                    if 'twitter.com' in href or 'x.com' in href:
                        social["x_handle"] = SocialMediaExtractor.extract_x(href)
                    elif 'instagram.com' in href:
                        social["instagram"] = SocialMediaExtractor.extract_instagram(href)
                    elif 'linkedin.com' in href:
                        social["linkedin"] = SocialMediaExtractor.extract_linkedin(href)
                except:
                    continue
    except:
        pass

    return social


async def extract_speaker(block: Locator, index: int, page: Page) -> Dict[str, str]:
    """Extract complete speaker information."""
    data = {
        "name": "",
        "tag": "",
        "image_url": "",
        "x_handle": "N/A",
        "linkedin": "N/A",
        "instagram": "N/A",
        "order": index  # Preserve exact website order
    }

    try:
        name_elem = block.locator("h3.sz-speaker__name").first
        if await name_elem.count():
            data["name"] = (await name_elem.inner_text()).strip()

        if not data["name"]:
            return data

        tag_elem = block.locator("h4.sz-speaker__tagline").first
        if await tag_elem.count():
            data["tag"] = (await tag_elem.inner_text()).strip()

        img = block.locator('img').first
        if await img.count():
            src = await img.get_attribute("src") or await img.get_attribute("data-src")
            if src:
                data["image_url"] = src if src.startswith('http') else f"https://www.nft.nyc{src}"

        # Click to open modal
        try:
            await block.click(force=True, timeout=2000)
            await page.wait_for_timeout(MODAL_WAIT_MS)
        except:
            return data

        # Extract from modal
        modal = page.locator('div.sz-speaker.sz-speaker--full').first
        if await modal.is_visible(timeout=1200):
            modal_name_elem = modal.locator('h3.sz-speaker__name').first
            if await modal_name_elem.count():
                modal_name = (await modal_name_elem.inner_text()).strip()
                if modal_name.lower() == data["name"].lower():
                    social = await extract_social_from_modal(modal)
                    data.update(social)

        await close_modal_completely(page)

    except Exception as e:
        logger.debug(f"Error extracting speaker: {e}")
        try:
            await close_modal_completely(page)
        except:
            pass

    return data


async def scrape_track(page: Page, track_name: str, track_url: str) -> List[Dict]:
    """Scrape speakers from a specific track page in EXACT website order."""

    full_url = f"{BASE_URL}{track_url}"
    logger.info(f"\n[{track_name}]")
    logger.info(f"  URL: {full_url}")

    try:
        # Navigate to track page
        await page.goto(full_url, wait_until='domcontentloaded', timeout=30000)

        # ⚡ IMPORTANT: Wait for page to fully render and stabilize
        await page.wait_for_timeout(PAGE_LOAD_WAIT_MS)

        # Wait for speaker grid to be visible
        await page.wait_for_selector('[data-speakerid]', state='visible', timeout=10000)

        await dismiss_cookie_banner(page)

        # Get speaker blocks - these should now be in DOM order
        speaker_blocks = page.locator('[data-speakerid]')
        count = await speaker_blocks.count()

        if count == 0:
            logger.warning(f"  No speakers found")
            return []

        logger.info(f"  Found {count} speakers (extracting in website order)")

        speakers = []

        with tqdm(total=count, desc=f"  Extracting", unit="speaker", leave=False) as pbar:
            for i in range(count):
                # ⚡ CRITICAL: Get fresh reference to maintain order
                blocks = page.locator('[data-speakerid]')
                block = blocks.nth(i)

                speaker = await extract_speaker(block, i, page)

                # Preserve exact position
                speaker['order'] = i
                speakers.append(speaker)
                pbar.update(1)

        with_x = sum(1 for s in speakers if s.get('x_handle') != 'N/A')
        with_ig = sum(1 for s in speakers if s.get('instagram') != 'N/A')
        with_li = sum(1 for s in speakers if s.get('linkedin') != 'N/A')

        # Log first speaker for verification
        if speakers:
            logger.info(f"  First speaker: {speakers[0]['name']}")
        logger.info(f"  ✓ {count} speakers | X:{with_x} IG:{with_ig} LI:{with_li}")

        return speakers

    except Exception as e:
        logger.error(f"  Error scraping {track_name}: {e}")
        return []


async def scrape_all_tracks(page: Page) -> Dict[str, List[Dict]]:
    """Scrape all track pages."""
    all_tracks_data = {}

    logger.info(f"\n{'=' * 60}")
    logger.info(f"SCRAPING {len(TRACKS)} TRACKS")
    logger.info(f"{'=' * 60}")

    for idx, (track_name, track_url) in enumerate(TRACKS, 1):
        logger.info(f"\n[Track {idx}/{len(TRACKS)}]")

        speakers = await scrape_track(page, track_name, track_url)

        if speakers:
            all_tracks_data[track_name] = speakers

        await page.wait_for_timeout(TRACK_PAUSE_MS)

    return all_tracks_data


def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Clean dataframe: remove duplicates, blanks, clean spacing - PRESERVE EXACT ORDER."""

    # Remove rows where name is empty
    df = df[df['name'].str.strip() != ''].copy()

    # Clean all string columns - remove extra whitespace
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].str.strip()
            # Replace multiple spaces with single space
            df[col] = df[col].str.replace(r'\s+', ' ', regex=True)

    # ⚡ CRITICAL: Sort by order FIRST to ensure website sequence
    df = df.sort_values('order', ascending=True)

    # Then remove duplicates, keeping first (earliest order) occurrence
    df = df.drop_duplicates(subset=['name'], keep='first')

    # Drop the order column (we don't need it in output)
    df = df.drop('order', axis=1)

    # Reset index to sequential numbers
    df = df.reset_index(drop=True)

    return df


def sanitize_sheet_name(name: str, used: Set[str]) -> str:
    """Sanitize Excel sheet name."""
    if not name:
        name = "Sheet"

    for char in [':', '\\', '/', '?', '*', '[', ']']:
        name = name.replace(char, '-')

    name = name.replace('&', 'and')

    if len(name) > 31:
        name = name[:31]

    base = name
    counter = 1
    while name in used:
        suffix = f"_{counter}"
        allowed = 31 - len(suffix)
        name = (base[:allowed] + suffix) if len(base) > allowed else (base + suffix)
        counter += 1

    used.add(name)
    return name


def apply_professional_formatting(excel_file: str, track_url_map: Dict[str, str]):
    """Apply professional Excel formatting to all sheets."""

    logger.info("\n📊 Applying professional formatting...")

    wb = load_workbook(excel_file)

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    alternate_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    na_font = Font(name='Calibri', size=11, italic=True, color="A6A6A6")
    normal_font = Font(name='Calibri', size=11)

    border_side = Side(style='medium', color='000000')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    metadata_font = Font(name='Calibri', size=10, italic=True, color="808080")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get the track URL for metadata
        track_name_original = sheet_name.replace(' and ', ' & ')
        track_url = track_url_map.get(track_name_original, "/speakers")
        source_url = f"{BASE_URL}{track_url}"

        max_row = ws.max_row
        max_col = ws.max_column

        # Format header row
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Format data rows
        for row in range(2, max_row + 1):
            # Alternate row shading
            if row % 2 == 0:
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).fill = alternate_fill

            # Apply borders and font styling
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border

                # Special formatting for "N/A" values
                if cell.value == "N/A":
                    cell.font = na_font
                else:
                    cell.font = normal_font

                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # Auto-fit column widths
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = ws['A2']

        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Add metadata at the bottom
        metadata_row = max_row + 2

        ws.cell(row=metadata_row, column=1).value = f"📊 Sourced from {source_url}"
        ws.cell(row=metadata_row, column=1).font = metadata_font

        timestamp = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        ws.cell(row=metadata_row + 1, column=1).value = f"⏰ Scraped on {timestamp}"
        ws.cell(row=metadata_row + 1, column=1).font = metadata_font

        logger.info(f"  ✓ Formatted sheet: {sheet_name}")

    wb.save(excel_file)
    logger.info(f"✓ Formatting complete!")


async def export_multi_track_data(all_tracks_data: Dict[str, List[Dict]], filename: str):
    """Export all tracks to Excel and JSON with professional formatting."""

    if not all_tracks_data:
        logger.error("No data to export")
        return

    excel_file = filename if filename.endswith('.xlsx') else f"{filename}.xlsx"
    used_sheet_names: Set[str] = set()

    logger.info("\n" + "=" * 60)
    logger.info("EXPORTING TO EXCEL")
    logger.info("=" * 60)

    track_url_map = {name: url for name, url in TRACKS}

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        for track_name, speakers in all_tracks_data.items():
            if not speakers:
                continue

            # Create DataFrame
            df = pd.DataFrame(
                speakers,
                columns=["name", "tag", "image_url", "x_handle", "linkedin", "instagram", "order"]
            )

            # Clean dataframe (preserves order)
            df = clean_dataframe(df)

            # Rename columns for display
            df.columns = ["Name", "Title/Tag", "Image URL", "X Handle", "LinkedIn", "Instagram"]

            sheet_name = sanitize_sheet_name(track_name, used_sheet_names)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Log first speaker for verification
            if len(df) > 0:
                logger.info(f"  ✓ Sheet '{sheet_name}': {len(df)} speakers (First: {df.iloc[0]['Name']})")
            else:
                logger.info(f"  ✓ Sheet '{sheet_name}': {len(df)} speakers")

    logger.info(f"\n✓ Saved Excel: {excel_file}")

    apply_professional_formatting(excel_file, track_url_map)

    # Save JSON
    json_file = excel_file.replace('.xlsx', '.json')

    total_speakers = sum(len(speakers) for speakers in all_tracks_data.values())
    total_with_x = sum(sum(1 for s in speakers if s.get('x_handle') != 'N/A') for speakers in all_tracks_data.values())
    total_with_ig = sum(
        sum(1 for s in speakers if s.get('instagram') != 'N/A') for speakers in all_tracks_data.values())
    total_with_li = sum(sum(1 for s in speakers if s.get('linkedin') != 'N/A') for speakers in all_tracks_data.values())

    json_data = {
        "metadata": {
            "scraped_at": datetime.now().isoformat(),
            "base_url": BASE_URL,
            "total_tracks": len(all_tracks_data),
            "total_speakers": total_speakers,
            "with_x": total_with_x,
            "with_instagram": total_with_ig,
            "with_linkedin": total_with_li,
        },
        "tracks": {}
    }

    for track_name, speakers in all_tracks_data.items():
        cleaned_speakers = []
        seen_names = set()

        # Sort by order to maintain exact sequence
        sorted_speakers = sorted(speakers, key=lambda x: x.get('order', 0))

        for speaker in sorted_speakers:
            if speaker['name'] and speaker['name'] not in seen_names:
                speaker_copy = {k: v for k, v in speaker.items() if k != 'order'}
                cleaned_speakers.append(speaker_copy)
                seen_names.add(speaker['name'])

        df = pd.DataFrame(cleaned_speakers, columns=["name", "tag", "image_url", "x_handle", "linkedin", "instagram"])

        with_x = sum(1 for s in cleaned_speakers if s.get('x_handle') != 'N/A')
        with_ig = sum(1 for s in cleaned_speakers if s.get('instagram') != 'N/A')
        with_li = sum(1 for s in cleaned_speakers if s.get('linkedin') != 'N/A')

        json_data["tracks"][track_name] = {
            "speaker_count": len(cleaned_speakers),
            "stats": {
                "with_x": with_x,
                "with_instagram": with_ig,
                "with_linkedin": with_li
            },
            "speakers": df.where(pd.notnull(df), None).to_dict('records')
        }

    with open(json_file, 'w') as f:
        json.dump(json_data, f, indent=2)

    logger.info(f"✓ Saved JSON: {json_file}")


async def main(args):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{BASE_FILENAME}_{timestamp}.xlsx" if USE_TIMESTAMP else f"{BASE_FILENAME}.xlsx"

    logger.info("=" * 60)
    logger.info("NFT NYC SPEAKER SCRAPER ⚡ OPTIMIZED")
    logger.info("=" * 60)
    logger.info(f"⚡ Performance: Modal {MODAL_WAIT_MS}ms | Close {MODAL_CLOSE_MS}ms | Load {PAGE_LOAD_WAIT_MS}ms")
    logger.info(f"📊 Tracks to scrape: {len(TRACKS)}")
    for i, (name, url) in enumerate(TRACKS, 1):
        logger.info(f"  {i:2d}. {name:20s} -> {url}")
    logger.info(f"\n📁 Output: {filename}")
    logger.info("=" * 60)

    start = time()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=args.headless)
        context = await browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = await context.new_page()

        all_tracks_data = await scrape_all_tracks(page)

        await browser.close()

    if all_tracks_data:
        await export_multi_track_data(all_tracks_data, filename)

        elapsed = time() - start

        # Count after cleaning
        total_cleaned = 0
        for speakers in all_tracks_data.values():
            seen = set()
            for s in sorted(speakers, key=lambda x: x.get('order', 0)):
                if s['name'] and s['name'] not in seen:
                    total_cleaned += 1
                    seen.add(s['name'])

        total_with_x = sum(sum(1 for s in speakers if s.get('x_handle') != 'N/A' and s.get('name')) for speakers in
                           all_tracks_data.values())
        total_with_ig = sum(sum(1 for s in speakers if s.get('instagram') != 'N/A' and s.get('name')) for speakers in
                            all_tracks_data.values())
        total_with_li = sum(sum(1 for s in speakers if s.get('linkedin') != 'N/A' and s.get('name')) for speakers in
                            all_tracks_data.values())

        logger.info("\n" + "=" * 60)
        logger.info("FINAL SUMMARY")
        logger.info("=" * 60)
        logger.info(f"⏱  Time: {elapsed:.1f}s ({elapsed / 60:.1f} minutes)")
        logger.info(f"⚡ Speed: {total_cleaned / elapsed:.1f} speakers/second")
        logger.info(f"📊 Tracks: {len(all_tracks_data)}/{len(TRACKS)}")
        logger.info(f"👥 Total Speakers (cleaned): {total_cleaned}")
        logger.info(f"🐦 X: {total_with_x}/{total_cleaned} ({total_with_x / max(total_cleaned, 1) * 100:.1f}%)")
        logger.info(f"📷 IG: {total_with_ig}/{total_cleaned} ({total_with_ig / max(total_cleaned, 1) * 100:.1f}%)")
        logger.info(f"💼 LI: {total_with_li}/{total_cleaned} ({total_with_li / max(total_cleaned, 1) * 100:.1f}%)")

        logger.info("\nBREAKDOWN BY TRACK:")
        logger.info("-" * 60)
        for track_name, speakers in all_tracks_data.items():
            seen = set()
            cleaned = []
            for s in sorted(speakers, key=lambda x: x.get('order', 0)):
                if s['name'] and s['name'] not in seen:
                    cleaned.append(s)
                    seen.add(s['name'])

            with_x = sum(1 for s in cleaned if s.get('x_handle') != 'N/A')
            with_ig = sum(1 for s in cleaned if s.get('instagram') != 'N/A')
            with_li = sum(1 for s in cleaned if s.get('linkedin') != 'N/A')

            # Show first speaker for verification
            first_name = cleaned[0]['name'] if cleaned else "N/A"
            logger.info(
                f"{track_name:20s} | {len(cleaned):3d} speakers | X:{with_x:2d} IG:{with_ig:2d} LI:{with_li:2d} | 1st: {first_name}")
        logger.info("=" * 60)
    else:
        logger.error("\n❌ No data extracted")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='NFT NYC Speaker Scraper ⚡ OPTIMIZED')
    parser.add_argument('--headless', action='store_true', help='Run in headless mode (faster)')
    args = parser.parse_args()

    try:
        asyncio.run(main(args))
    except KeyboardInterrupt:
        logger.info("\n\n⚠️  Interrupted by user")
    except Exception as e:
        logger.error(f"\n❌ Fatal error: {e}")
        import traceback

        traceback.print_exc()